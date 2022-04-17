from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import urllib3
import io
import PIL

wb=Workbook()
ws=wb.active



#for row in range(1,5):
#    for col in range(1,5):
#        char=get_column_letter(col)
#        print(ws[char+str(row)].value)

#ws.merge_cells('A1:C3')
#ws.unmerge_cells("A1:C3")

#ws.move_range("C1:D4",rows=2,cols=2)

for col in range(1,6):
    ws[get_column_letter(col)+"1"].font=Font(bold=True,color='0099CCFF')

ws.column_dimensions["C"].width=5.71
#ws.row_dimensions[3].height=20

r = 1
ws['A1'] = 'test'
#Client HTTP
http = urllib3.PoolManager()
r = http.request('GET', 'http://ddragon.leagueoflegends.com/cdn/11.16.1/img/champion/Aatrox.png')
image_file = io.BytesIO(r.data)
img = Image(image_file)

for row in range(100): #Ligne
    ws.row_dimensions[row+1].height=30


img.height = 40# insert image height in pixels as float or int (e.g. 305.5)
img.width=40 # insert image width in pixels as float or int (e.g. 405.8)
img.anchor = 'C1' # where you want image to be anchored/start from

ws.add_image(img)

wb.save('test.xlsx')