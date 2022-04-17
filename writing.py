import requete as rq
import main as m
import csv
#Importation excel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import urllib3
import io
import PIL

def excel():
    tab=[]
    en=[]
    wb=Workbook()
    ws=wb.active
    for i in m.matchdata():
        ws.append(i)
    
    #Dimension des cases
    ws.column_dimensions["A"].width=5.71
    ws.column_dimensions["C"].width=5.71
    for row in range(len(m.matchdata())): #Ligne
        ws.row_dimensions[row+1].height=30
    #Mettre les champions joué dans un tableau
    for i in m.matchdata():
        tab.append(i[0])
    
    for k in m.matchdata():
        en.append(k[2])
    
    for e in range(1,len(en)):
        http = urllib3.PoolManager()
        r = http.request('GET', f'http://ddragon.leagueoflegends.com/cdn/11.16.1/img/champion/{en[e]}.png')
        image_file = io.BytesIO(r.data)
        img = Image(image_file)
        img.height = 40
        img.width=40
        img.anchor = 'C'+str(e+1)
        ws.add_image(img)

    for j in range(1,len(tab)): #Setup l'image
        http = urllib3.PoolManager()
        r = http.request('GET', f'http://ddragon.leagueoflegends.com/cdn/11.16.1/img/champion/{tab[j]}.png')
        image_file = io.BytesIO(r.data)
        img = Image(image_file)
        img.height = 40
        img.width=40 
        img.anchor = 'A'+str(j+1)
        ws.add_image(img)

    wb.save(f"{rq.nametmp}.xlsx")
    print("ECRITURE TERMINÉE")

def csv_fct():
    with open(f"data_out/match_{rq.nametmp}.csv","w") as csvfile:
        writer=csv.writer(csvfile)
        writer.writerows(m.matchdata())
    print("ECRITURE EXCEL TERMINÉ")

def txt():
    f = open(f"data_out/{rq.nametmp}.txt",'w')
    tmp=m.profil()
    for i in tmp:
        if type(i)==list:
            for j in range(len(i)):
                f.write(str(i[j])+"\n")
            else:
                continue
        f.write(str(i)+"\n")
        f.write("\n")
    f.close()
    print("ECRITURE TXT TERMINÉ")

excel()